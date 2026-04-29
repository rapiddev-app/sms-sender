"""Чтение и валидация Excel-файла с контактами для SMS-рассылки.

Структура входного файла:
    столбец A — номер телефона
    столбец B — переменная для подстановки в шаблон
    первая строка — заголовок (пропускается)

Модуль не зависит от GUI и ADB-слоя — это чистая бизнес-логика.
"""

import re
from pathlib import Path

from openpyxl import load_workbook

from core.models import Contact, LoadResult, ValidationError

_HEADER_ROW = 1
_PHONE_CLEANUP_RE = re.compile(r"[ \-()\.\+]")


class InvalidExcelFormatError(Exception):
    """Файл существует, но не является валидным `.xlsx` или повреждён."""


def load_excel(path: str | Path) -> LoadResult:
    """Загружает контакты из `.xlsx`.

    Возвращает структурированный результат: валидные контакты и список ошибок
    с привязкой к номеру строки. Битые данные внутри файла наружу исключения
    не бросают — они попадают в `LoadResult.errors`.

    Raises:
        FileNotFoundError: файл по указанному пути не существует.
        InvalidExcelFormatError: файл не открывается как `.xlsx`.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidExcelFormatError(
            f"Не удалось прочитать файл как .xlsx: {path.name}"
        ) from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            return LoadResult(contacts=[], errors=[])

        contacts: list[Contact] = []
        errors: list[ValidationError] = []
        seen_phones: set[str] = set()

        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_idx == _HEADER_ROW:
                continue

            raw_phone = row[0] if len(row) > 0 else None
            raw_variable = row[1] if len(row) > 1 else None

            # Хвостовые пустые строки в Excel — пропускаем без ошибки.
            if _is_empty(raw_phone) and _is_empty(raw_variable):
                continue

            phone_display = _to_str(raw_phone)
            variable_display = _to_str(raw_variable)

            if _is_empty(raw_phone):
                errors.append(
                    ValidationError(
                        row=row_idx,
                        raw_phone=phone_display,
                        raw_variable=variable_display,
                        reason="empty_phone",
                    )
                )
                continue

            normalized = normalize_phone(raw_phone)
            if normalized is None:
                errors.append(
                    ValidationError(
                        row=row_idx,
                        raw_phone=phone_display,
                        raw_variable=variable_display,
                        reason="invalid_phone",
                    )
                )
                continue

            if _is_empty(raw_variable):
                errors.append(
                    ValidationError(
                        row=row_idx,
                        raw_phone=phone_display,
                        raw_variable=variable_display,
                        reason="empty_variable",
                    )
                )
                continue

            if normalized in seen_phones:
                errors.append(
                    ValidationError(
                        row=row_idx,
                        raw_phone=phone_display,
                        raw_variable=variable_display,
                        reason="duplicate",
                    )
                )
                continue

            seen_phones.add(normalized)
            contacts.append(
                Contact(
                    row=row_idx,
                    phone=normalized,
                    variable=variable_display.strip(),
                )
            )

        return LoadResult(contacts=contacts, errors=errors)
    finally:
        workbook.close()


def normalize_phone(raw: object) -> str | None:
    """Нормализует номер к формату `+7xxxxxxxxxx`.

    Поддерживается: 10 цифр (мобильный без префикса, старт с 9),
    11 цифр со старта 7 или 8, формат `+7xxxxxxxxxx`. Городские номера
    отбрасываются — мобильный код всегда начинается с 9.

    Возвращает `None`, если формат невалиден.
    """
    if raw is None:
        return None

    if isinstance(raw, bool):
        return None

    if isinstance(raw, (int, float)):
        # Excel часто хранит номера как числа — приводим к строке без ".0".
        if isinstance(raw, float) and not raw.is_integer():
            return None
        digits = str(int(raw))
    else:
        digits = _PHONE_CLEANUP_RE.sub("", str(raw))

    if not digits or not digits.isdigit():
        return None

    if len(digits) == 10 and digits.startswith("9"):
        return "+7" + digits
    if len(digits) == 11 and digits[0] in ("7", "8") and digits[1] == "9":
        return "+7" + digits[1:]
    return None


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)
