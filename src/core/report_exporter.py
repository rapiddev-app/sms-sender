"""Экспорт результата рассылки обратно в Excel-файл."""

from pathlib import Path
from typing import Final

from openpyxl import load_workbook

STATUS_HEADER: Final = "Статус"
STATUS_PENDING: Final = "⏳"
STATUS_SENT: Final = "✅"
STATUS_FAILED: Final = "❌"

_HEADER_ROW: Final = 1


class ReportExportError(Exception):
    """Ошибка экспорта отчёта в `.xlsx`."""


def export_status_report(
    source_path: str | Path,
    destination_path: str | Path,
    statuses_by_row: dict[int, str],
) -> None:
    """Сохраняет копию Excel-файла со столбцом `Статус`.

    `statuses_by_row` использует номера строк исходного Excel-файла.

    Raises:
        FileNotFoundError: исходный файл не найден.
        ReportExportError: файл не удалось прочитать или сохранить.
    """
    source = Path(source_path)
    destination = Path(destination_path)
    if not source.exists():
        raise FileNotFoundError(f"Файл не найден: {source}")

    try:
        workbook = load_workbook(filename=source)
    except Exception as exc:
        raise ReportExportError(f"Не удалось прочитать Excel-файл: {source.name}") from exc

    try:
        worksheet = workbook.active
        status_column = _find_or_create_status_column(worksheet)
        for row_index, status in sorted(statuses_by_row.items()):
            if row_index <= _HEADER_ROW:
                continue
            worksheet.cell(row=row_index, column=status_column, value=status)

        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
    except OSError as exc:
        raise ReportExportError(f"Не удалось сохранить отчёт: {destination}") from exc
    finally:
        workbook.close()


def _find_or_create_status_column(worksheet) -> int:
    for column_index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=_HEADER_ROW, column=column_index).value
        if value == STATUS_HEADER:
            return column_index

    status_column = worksheet.max_column + 1
    worksheet.cell(row=_HEADER_ROW, column=status_column, value=STATUS_HEADER)
    return status_column
