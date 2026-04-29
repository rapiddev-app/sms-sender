import pytest
from openpyxl import Workbook, load_workbook

from core.report_exporter import (
    STATUS_FAILED,
    STATUS_HEADER,
    STATUS_PENDING,
    STATUS_SENT,
    ReportExportError,
    export_status_report,
)


def _create_contacts_workbook(path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Телефон", "Переменная"])
    worksheet.append(["+79990000001", "Анна"])
    worksheet.append(["+79990000002", "Иван"])
    worksheet.append(["+79990000003", "Ольга"])
    workbook.save(path)
    workbook.close()


def _read_rows(path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path)
    try:
        worksheet = workbook.active
        return [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def test_export_status_report_adds_status_column(tmp_path) -> None:
    source_path = tmp_path / "contacts.xlsx"
    destination_path = tmp_path / "report.xlsx"
    _create_contacts_workbook(source_path)

    export_status_report(
        source_path,
        destination_path,
        {
            2: STATUS_SENT,
            3: STATUS_FAILED,
            4: STATUS_PENDING,
        },
    )

    assert _read_rows(destination_path) == [
        ("Телефон", "Переменная", STATUS_HEADER),
        ("+79990000001", "Анна", STATUS_SENT),
        ("+79990000002", "Иван", STATUS_FAILED),
        ("+79990000003", "Ольга", STATUS_PENDING),
    ]


def test_export_status_report_reuses_existing_status_column(tmp_path) -> None:
    source_path = tmp_path / "contacts.xlsx"
    destination_path = tmp_path / "report.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Телефон", STATUS_HEADER, "Переменная"])
    worksheet.append(["+79990000001", "старый", "Анна"])
    workbook.save(source_path)
    workbook.close()

    export_status_report(source_path, destination_path, {2: STATUS_SENT})

    assert _read_rows(destination_path) == [
        ("Телефон", STATUS_HEADER, "Переменная"),
        ("+79990000001", STATUS_SENT, "Анна"),
    ]


def test_export_status_report_ignores_header_row_status(tmp_path) -> None:
    source_path = tmp_path / "contacts.xlsx"
    destination_path = tmp_path / "report.xlsx"
    _create_contacts_workbook(source_path)

    export_status_report(source_path, destination_path, {1: STATUS_FAILED, 2: STATUS_SENT})

    assert _read_rows(destination_path)[0] == ("Телефон", "Переменная", STATUS_HEADER)


def test_export_status_report_raises_for_missing_source(tmp_path) -> None:
    with pytest.raises(FileNotFoundError):
        export_status_report(tmp_path / "missing.xlsx", tmp_path / "report.xlsx", {})


def test_export_status_report_raises_for_invalid_xlsx(tmp_path) -> None:
    source_path = tmp_path / "broken.xlsx"
    source_path.write_text("not xlsx", encoding="utf-8")

    with pytest.raises(ReportExportError, match="Не удалось прочитать Excel-файл"):
        export_status_report(source_path, tmp_path / "report.xlsx", {})
